"""
Excel-Export-Modul des Controlling-Tools.

Erstellt einen professionellen Excel-Bericht mit 4 Sheets:
  1. Zusammenfassung  – KPI-Übersicht mit Ampelfarben
  2. Plan-Ist-Vergleich – Detailtabelle mit bedingter Formatierung
  3. Zeitreihe        – Monatliche Entwicklung mit eingebettetem Diagramm
  4. Forecast         – Jahreshochrechnung

Verwendet openpyxl für alle Formatierungen.
"""

import datetime
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import config
from src.transformer import (
    aggregate_by_month,
    calculate_full_year_forecast,
    calculate_plan_ist_comparison,
    calculate_ytd,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Farb-Konstanten (Hex ohne #)
# ---------------------------------------------------------------------------
FARBE_HEADER_BG = "1F3864"      # Dunkelblau
FARBE_HEADER_FG = "FFFFFF"      # Weiß
FARBE_ZEILE_GERADE = "EEF2FF"   # Sehr helles Blau
FARBE_ZEILE_UNGERADE = "FFFFFF"  # Weiß
FARBE_GRUEN = "C6EFCE"          # Grün (Excel-Stil)
FARBE_GELB = "FFEB9C"           # Gelb
FARBE_ROT = "FFC7CE"            # Rot
FARBE_TITEL_BG = "2E75B6"       # Mittelblau für Titel-Zellen

# Währungsformat mit Tausendertrennzeichen
FORMAT_WAEHRUNG = f'#,##0.00 "{config.CURRENCY_SYMBOL}"'
FORMAT_PROZENT = '+0.00%;-0.00%;0.00%'
FORMAT_ZAHL = '#,##0.00'


# ---------------------------------------------------------------------------
# Hilfsfunktionen für Formatierung
# ---------------------------------------------------------------------------

def _header_fill(farbe: str = FARBE_HEADER_BG) -> PatternFill:
    return PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")

def _zeilen_fill(farbe: str) -> PatternFill:
    return PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")

def _dünner_rahmen() -> Border:
    seite = Side(style="thin", color="BFBFBF")
    return Border(left=seite, right=seite, top=seite, bottom=seite)

def _formatiere_header_zeile(ws, zeile: int, anzahl_spalten: int) -> None:
    """Formatiert eine Kopfzeile (dunkelblau, weiß, fett)."""
    for col in range(1, anzahl_spalten + 1):
        zelle = ws.cell(row=zeile, column=col)
        zelle.fill = _header_fill()
        zelle.font = Font(bold=True, color=FARBE_HEADER_FG, size=10)
        zelle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        zelle.border = _dünner_rahmen()

def _formatiere_datenzelle(ws, zeile: int, col: int, gerade: bool) -> None:
    """Formatiert eine normale Datenzelle mit abwechselnden Zeilen."""
    zelle = ws.cell(row=zeile, column=col)
    farbe = FARBE_ZEILE_GERADE if gerade else FARBE_ZEILE_UNGERADE
    zelle.fill = _zeilen_fill(farbe)
    zelle.border = _dünner_rahmen()
    zelle.alignment = Alignment(vertical="center")

def _abweichungsfarbe(abw_pct: float) -> str:
    """Gibt die Hintergrundfarbe für eine Abweichungszelle zurück."""
    if pd.isna(abw_pct):
        return FARBE_ZEILE_UNGERADE
    abs_abw = abs(abw_pct)
    if abs_abw < config.TRAFFIC_LIGHT_GREEN_THRESHOLD:
        return FARBE_GRUEN
    elif abs_abw < config.TRAFFIC_LIGHT_YELLOW_THRESHOLD:
        return FARBE_GELB
    else:
        return FARBE_ROT

def _spaltenbreite_anpassen(ws, min_breite: int = 12, max_breite: int = 30) -> None:
    """Passt alle Spaltenbreiten automatisch an den Inhalt an."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for zelle in col:
            try:
                if zelle.value:
                    max_len = max(max_len, len(str(zelle.value)))
            except Exception:
                pass
        breite = max(min_breite, min(max_len + 2, max_breite))
        ws.column_dimensions[col_letter].width = breite


# ---------------------------------------------------------------------------
# Sheet-Erstellungsfunktionen
# ---------------------------------------------------------------------------

def _erstelle_zusammenfassung(ws, ytd_df: pd.DataFrame, forecast_df: pd.DataFrame,
                               stichtag: datetime.date) -> None:
    """Sheet 1: KPI-Übersicht."""
    ws.title = "Zusammenfassung"

    # Titel
    ws.merge_cells("A1:F1")
    titel = ws["A1"]
    titel.value = f"Controlling-Übersicht | Stichtag: {stichtag.strftime('%d.%m.%Y')}"
    titel.font = Font(bold=True, size=14, color=FARBE_HEADER_FG)
    titel.fill = _header_fill(FARBE_TITEL_BG)
    titel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # KPI-Tabelle YTD
    ws["A3"] = "YTD-Kennzahlen (Year-to-Date)"
    ws["A3"].font = Font(bold=True, size=11)

    kpi_header = ["Kostenstelle", "Kategorie", "YTD Budget (€)", "YTD Ist (€)",
                  "Abweichung (€)", "Abweichung (%)", "Ampel"]
    for col_idx, h in enumerate(kpi_header, start=1):
        ws.cell(row=4, column=col_idx).value = h
    _formatiere_header_zeile(ws, 4, len(kpi_header))

    for i, (_, row) in enumerate(ytd_df.iterrows(), start=5):
        gerade = (i % 2 == 0)
        werte = [
            row.get("Kostenstelle", ""),
            row.get("Kategorie", ""),
            row.get("YTD_Budget", 0),
            row.get("YTD_Ist", 0),
            row.get("Abweichung_absolut", 0),
            row.get("Abweichung_prozent", 0) / 100,
            row.get("Ampel", ""),
        ]
        for col_idx, wert in enumerate(werte, start=1):
            zelle = ws.cell(row=i, column=col_idx, value=wert)
            _formatiere_datenzelle(ws, i, col_idx, gerade)
            if col_idx in (3, 4, 5):
                zelle.number_format = FORMAT_WAEHRUNG
            elif col_idx == 6:
                zelle.number_format = FORMAT_PROZENT
                abw = row.get("Abweichung_prozent", 0)
                zelle.fill = _zeilen_fill(_abweichungsfarbe(abw))
            elif col_idx == 7:
                zelle.alignment = Alignment(horizontal="center", vertical="center")

    _spaltenbreite_anpassen(ws)


def _erstelle_plan_ist(ws, comparison_df: pd.DataFrame) -> None:
    """Sheet 2: Detaillierter Plan/Ist-Vergleich."""
    ws.title = "Plan-Ist-Vergleich"

    spalten = ["Jahr", "Monat", "Kostenstelle", "Kategorie",
               "Budget", "Ist", "Abweichung_absolut", "Abweichung_prozent", "Ampel"]
    vorhandene = [s for s in spalten if s in comparison_df.columns]
    df_export = comparison_df[vorhandene].copy()

    # Monatsname ergänzen
    monatsnamen = {1: "Jan", 2: "Feb", 3: "Mär", 4: "Apr", 5: "Mai", 6: "Jun",
                   7: "Jul", 8: "Aug", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dez"}
    df_export.insert(2, "Monatsname", df_export["Monat"].map(monatsnamen))

    # Header
    header = list(df_export.columns)
    for col_idx, h in enumerate(header, start=1):
        ws.cell(row=1, column=col_idx).value = h
    _formatiere_header_zeile(ws, 1, len(header))
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Daten
    for i, (_, row) in enumerate(df_export.iterrows(), start=2):
        gerade = (i % 2 == 0)
        for col_idx, (col_name, wert) in enumerate(row.items(), start=1):
            zelle = ws.cell(row=i, column=col_idx, value=wert)
            _formatiere_datenzelle(ws, i, col_idx, gerade)

            if col_name in ("Budget", "Ist", "Abweichung_absolut"):
                zelle.number_format = FORMAT_WAEHRUNG
            elif col_name == "Abweichung_prozent":
                zelle.number_format = FORMAT_PROZENT
                zelle.value = (wert or 0) / 100
                zelle.fill = _zeilen_fill(_abweichungsfarbe(wert or 0))
            elif col_name == "Ampel":
                zelle.alignment = Alignment(horizontal="center")

    _spaltenbreite_anpassen(ws)


def _erstelle_zeitreihe(ws, monthly_df: pd.DataFrame) -> None:
    """Sheet 3: Monatliche Zeitreihe mit eingebettetem Balkendiagramm."""
    ws.title = "Zeitreihe"

    header = ["Jahr", "Monat", "Budget (€)", "Ist (€)", "Abweichung (€)", "Abw. (%)"]
    for col_idx, h in enumerate(header, start=1):
        ws.cell(row=1, column=col_idx).value = h
    _formatiere_header_zeile(ws, 1, len(header))

    for i, (_, row) in enumerate(monthly_df.iterrows(), start=2):
        gerade = (i % 2 == 0)
        werte = [
            row.get("Jahr", ""),
            row.get("Monat", ""),
            row.get("Budget", 0),
            row.get("Ist", 0),
            row.get("Abweichung_absolut", 0),
            row.get("Abweichung_prozent", 0) / 100,
        ]
        for col_idx, wert in enumerate(werte, start=1):
            zelle = ws.cell(row=i, column=col_idx, value=wert)
            _formatiere_datenzelle(ws, i, col_idx, gerade)
            if col_idx in (3, 4, 5):
                zelle.number_format = FORMAT_WAEHRUNG
            elif col_idx == 6:
                zelle.number_format = FORMAT_PROZENT

    # Eingebettetes Balkendiagramm
    anzahl_zeilen = len(monthly_df) + 1
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Monatlicher Plan/Ist-Vergleich"
    chart.y_axis.title = f"Betrag ({config.CURRENCY_SYMBOL})"
    chart.x_axis.title = "Monat"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    budget_ref = Reference(ws, min_col=3, min_row=1, max_row=anzahl_zeilen)
    ist_ref = Reference(ws, min_col=4, min_row=1, max_row=anzahl_zeilen)
    chart.add_data(budget_ref, titles_from_data=True)
    chart.add_data(ist_ref, titles_from_data=True)

    ws.add_chart(chart, "H2")
    _spaltenbreite_anpassen(ws)


def _erstelle_forecast(ws, forecast_df: pd.DataFrame, stichtag: datetime.date) -> None:
    """Sheet 4: Jahreshochrechnung (Forecast)."""
    ws.title = "Forecast"

    ws.merge_cells("A1:G1")
    titel = ws["A1"]
    titel.value = f"Jahres-Forecast | Hochrechnung per {stichtag.strftime('%d.%m.%Y')}"
    titel.font = Font(bold=True, size=12, color=FARBE_HEADER_FG)
    titel.fill = _header_fill(FARBE_TITEL_BG)
    titel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    spalten_map = {
        "Kostenstelle": "Kostenstelle",
        "Kategorie": "Kategorie",
        "Budget_Gesamtjahr": "Budget Gesamtjahr (€)",
        "YTD_Ist": "YTD Ist (€)",
        "Restbudget": "Restbudget (€)",
        "Forecast_Gesamtjahr": "Forecast (€)",
        "Forecast_Abweichung_absolut": "Abweichung (€)",
        "Forecast_Abweichung_prozent": "Abweichung (%)",
        "Ampel": "Ampel",
    }

    vorhandene_spalten = [k for k in spalten_map if k in forecast_df.columns]
    header_labels = [spalten_map[k] for k in vorhandene_spalten]

    for col_idx, h in enumerate(header_labels, start=1):
        ws.cell(row=2, column=col_idx).value = h
    _formatiere_header_zeile(ws, 2, len(header_labels))

    waehrungs_spalten = {
        "Budget_Gesamtjahr", "YTD_Ist", "Restbudget",
        "Forecast_Gesamtjahr", "Forecast_Abweichung_absolut"
    }

    for i, (_, row) in enumerate(forecast_df[vorhandene_spalten].iterrows(), start=3):
        gerade = (i % 2 == 0)
        for col_idx, col_name in enumerate(vorhandene_spalten, start=1):
            wert = row[col_name]
            zelle = ws.cell(row=i, column=col_idx, value=wert)
            _formatiere_datenzelle(ws, i, col_idx, gerade)
            if col_name in waehrungs_spalten:
                zelle.number_format = FORMAT_WAEHRUNG
            elif col_name == "Forecast_Abweichung_prozent":
                zelle.number_format = FORMAT_PROZENT
                zelle.value = (wert or 0) / 100
                zelle.fill = _zeilen_fill(_abweichungsfarbe(wert or 0))
            elif col_name == "Ampel":
                zelle.alignment = Alignment(horizontal="center")

    _spaltenbreite_anpassen(ws)


# ---------------------------------------------------------------------------
# Hauptexport-Funktion
# ---------------------------------------------------------------------------
def export_to_excel(
    budget_df: pd.DataFrame,
    actuals_df: pd.DataFrame,
    as_of_date: Optional[datetime.date] = None,
    output_path: Optional[str] = None,
) -> Path:
    """Erstellt den vollständigen Excel-Bericht mit 4 Sheets.

    Args:
        budget_df:    Budget-DataFrame.
        actuals_df:   Ist-DataFrame.
        as_of_date:   Stichtag; Standard: heutiges Datum.
        output_path:  Ausgabepfad; Standard: config.DEFAULT_EXCEL_OUTPUT.

    Returns:
        Path-Objekt des erstellten Excel-Berichts.
    """
    stichtag = as_of_date or config.DEFAULT_AS_OF_DATE
    ausgabe = Path(output_path) if output_path else config.DEFAULT_EXCEL_OUTPUT
    ausgabe.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Erstelle Excel-Bericht: {ausgabe}")

    # Daten berechnen
    comparison_df = calculate_plan_ist_comparison(budget_df, actuals_df, stichtag)
    ytd_df = calculate_ytd(budget_df, actuals_df, stichtag)
    forecast_df = calculate_full_year_forecast(budget_df, actuals_df, stichtag)
    monthly_df = aggregate_by_month(comparison_df)

    # Workbook erstellen
    wb = Workbook()
    wb.remove(wb.active)  # Standard-Sheet entfernen

    _erstelle_zusammenfassung(wb.create_sheet(), ytd_df, forecast_df, stichtag)
    _erstelle_plan_ist(wb.create_sheet(), comparison_df)
    _erstelle_zeitreihe(wb.create_sheet(), monthly_df)
    _erstelle_forecast(wb.create_sheet(), forecast_df, stichtag)

    wb.save(ausgabe)
    logger.info(f"Excel-Bericht gespeichert: {ausgabe}")
    return ausgabe